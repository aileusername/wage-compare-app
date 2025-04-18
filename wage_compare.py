import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def extract_data(file_path):
    """Extract wage data from a structured text file."""
    codes, wd_dates, jobs, subclasses, rates, fringes = [], [], [], [], [], []

    header_pattern = re.compile(r'^\s*(?:\*\s*)?([A-Z0-9]+-\d{3})\s+(\d{2}/\d{2}/\d{4})\s*$')
    dot_line_pattern = re.compile(r'^(.*?)\.{2,}.*?\$\s*([\d,.]+\.\d{2})[^\d]*([\d,.]+\.\d{2})')
    group_title_pattern = re.compile(r'^[A-Z][A-Z \-/:()\']+$')
    separator_pattern = re.compile(r'^[-=]{3,}$')

    current_code = ""
    current_date = ""
    current_group = None
    pending_job_line = ""

    def clean_title(raw_title):
        for junk in ["Rates Fringes", "Rates", "Fringes"]:
            if raw_title.strip().upper().startswith(junk.upper()):
                raw_title = raw_title[len(junk):].strip()
        return raw_title.strip()

    def split_job_and_subclass(title):
        if '(' in title and title.endswith(')'):
            i = title.index('(')
            return title[:i].strip(), title[i:].strip()
        return title.strip(), ""

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if header_match := header_pattern.match(line):
                current_code = header_match.group(1)
                current_date = header_match.group(2)
                current_group = None
                pending_job_line = ""
                continue

            if separator_pattern.match(line):
                current_group = None
                pending_job_line = ""
                continue

            if group_title_pattern.match(line) and '$' not in line and '.' not in line:
                current_group = line.strip()
                pending_job_line = ""
                continue

            if '$' in line and '..' in line:
                full_title = f"{pending_job_line} {line}".strip() if pending_job_line else line
                pending_job_line = ""

                if dot_match := dot_line_pattern.match(full_title):
                    raw_title = dot_match.group(1).strip()
                    rate = dot_match.group(2).replace(',', '')
                    fringe = dot_match.group(3).replace(',', '')

                    clean = clean_title(raw_title)

                    if current_group:
                        if clean.upper() == current_group.upper():
                            job = clean
                            subclass = ""
                        else:
                            job = current_group
                            subclass = clean
                    else:
                        job, subclass = split_job_and_subclass(clean)

                    codes.append(current_code)
                    wd_dates.append(current_date)
                    jobs.append(job)
                    subclasses.append(subclass)
                    rates.append(rate)
                    fringes.append(fringe)
                continue

            if not ('$' in line or '..' in line or group_title_pattern.match(line)):
                pending_job_line = line.strip()

    return pd.DataFrame({
        'Code': codes,
        'WD_Date': wd_dates,
        'Job': jobs,
        'Job_Subclass': subclasses,
        'Rate': rates,
        'Fringe': fringes
    })


def compare_variants(df1, df2, rev1_label="v1", rev2_label="v2"):
    """Compare wage data and return added/removed/modified rows with labeled columns."""
    key_cols = ['Job', 'Job_Subclass']
    df1 = df1.copy()
    df2 = df2.copy()
    df1['__source__'] = rev1_label
    df2['__source__'] = rev2_label

    merged = df1.merge(df2, how="outer", on=key_cols, suffixes=('_1', '_2'), indicator=True)
    changes = []

    for _, row in merged.iterrows():
        change = ''
        if row['_merge'] == 'left_only':
            change = 'Removed'
        elif row['_merge'] == 'right_only':
            change = 'Added'
        elif row['Rate_1'] != row['Rate_2'] or row['Fringe_1'] != row['Fringe_2']:
            change = 'Modified'

        if change:
            changes.append({
                'Job': row['Job'],
                'Job_Subclass': row['Job_Subclass'],
                'Change_Type': change,
                f'Rate_{rev1_label}': row.get('Rate_1', ''),
                f'Rate_{rev2_label}': row.get('Rate_2', ''),
                f'Fringe_{rev1_label}': row.get('Fringe_1', ''),
                f'Fringe_{rev2_label}': row.get('Fringe_2', ''),
            })

    return pd.DataFrame(changes)


def apply_excel_styling(file_path):
    """Style Excel output: bold headers, currency formatting, auto-width."""
    wb = load_workbook(file_path)
    currency_format = '"$"#,##0.00'
    bold = Font(bold=True)

    for ws in wb.worksheets:
        if ws.max_row <= 1:
            continue

        for cell in ws[1]:
            cell.font = bold

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        currency_cols = [name for name in headers if 'Rate' in name or 'Fringe' in name]

        for row in range(2, ws.max_row + 1):
            for col_name in currency_cols:
                col_idx = headers[col_name]
                cell = ws.cell(row=row, column=col_idx)
                try:
                    cell.value = float(cell.value)
                    cell.number_format = currency_format
                except (ValueError, TypeError):
                    continue

    wb.save(file_path)


def get_rev_label(file_path, fallback="v1"):
    """Extract 'r0', 'r1', etc. from file name, else fallback."""
    match = re.search(r'\.(r\d+)\.txt$', os.path.basename(file_path), re.IGNORECASE)
    return match.group(1) if match else fallback